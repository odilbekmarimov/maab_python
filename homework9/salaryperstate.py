import sqlite3
import pandas as pd
from openpyxl import load_workbook

with sqlite3.connect('population.db') as con:
    df = pd.read_sql("select * from population", con = con)

uniqueStates = df.groupby('state')
StatePopulation = uniqueStates.size()
AvgSalaryState = uniqueStates['salary'].mean()
MedianSalaryState = uniqueStates['salary'].median()

populationTotal = len(df)
populationPercent = (StatePopulation / populationTotal) * 100

analysis_df = pd.DataFrame({
    'State': StatePopulation.index,
    'Percentage': populationPercent.values,
    'Average Salary': AvgSalaryState.values,
    'Median Salary': MedianSalaryState.values,
    'Number of population': StatePopulation.values
})

file_path = "population salary analysis.xlsx"
try:
    workbook = load_workbook(file_path)
    if "State" not in workbook.sheetnames:
        writer = pd.ExcelWriter(file_path, engine="openpyxl", mode="a")
        analysis_df.to_excel(writer, sheet_name="State", index=False, startrow=1, header=False)
        writer.close()
    else:
        with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            analysis_df.to_excel(writer, sheet_name="State", index=False, startrow=1, header=False)
            
except FileNotFoundError:
    headers = ["State", "Percentage", "Average Salary", "Median Salary", "Number of population"]
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        analysis_df.to_excel(writer, sheet_name="State", index=False, startrow=1, header=headers)

print(f"Data exported to '{file_path}'")
